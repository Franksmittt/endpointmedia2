#!/usr/bin/env python3
"""Export Google Search Console spreadsheet (all tabs) to formatted TXT."""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX_PATH = Path('/tmp/sheet.xlsx')
OUTPUT_PATH = Path('/workspace/GSC_SPREADSHEET_REPORT.txt')
SOURCE_URL = (
    'https://docs.google.com/spreadsheets/d/'
    '1lAlUZOwoeY03R8QknfqD_3PAapZDFQV_96IMQEh3DL4/edit?usp=sharing'
)


def fmt(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e12:
            if 0 < abs(val) < 1:
                return f'{val:.4f}'.rstrip('0').rstrip('.')
            return str(int(val)) if val == int(val) else f'{val:.2f}'
        return f'{val:.2f}'
    return str(val)


def pct(val):
    if val is None or val == '':
        return ''
    try:
        return f'{float(val) * 100:.2f}%'
    except (TypeError, ValueError):
        return fmt(val)


def write_table(lines, rows, header):
    col_widths = [len(h) for h in header]
    data_rows = []

    for row in rows[1:]:
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        formatted = []
        for i, cell in enumerate(row):
            if i < len(header) and header[i].upper() == 'CTR':
                formatted.append(pct(cell) if cell not in (None, '') else '')
            else:
                formatted.append(fmt(cell))
        data_rows.append(formatted)
        for i, value in enumerate(formatted):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(value))

    sep = ' | '
    header_line = sep.join(h.ljust(col_widths[i]) for i, h in enumerate(header))
    lines.append(header_line)
    lines.append('-' * len(header_line))

    for formatted in data_rows:
        line = sep.join(
            formatted[i].ljust(col_widths[i]) if i < len(col_widths) else formatted[i]
            for i in range(len(formatted))
        )
        lines.append(line)

    return data_rows


def add_chart_summary(lines, data_rows):
    if not data_rows:
        return
    total_clicks = sum(float(r[1]) for r in data_rows if r[1])
    total_impr = sum(float(r[2]) for r in data_rows if r[2])
    avg_pos_vals = [float(r[4]) for r in data_rows if r[4]]
    avg_pos = sum(avg_pos_vals) / len(avg_pos_vals) if avg_pos_vals else 0
    lines.append('')
    lines.append('SUMMARY (Chart / Daily):')
    lines.append(f'  Total days: {len(data_rows)}')
    lines.append(f'  Total clicks: {int(total_clicks)}')
    lines.append(f'  Total impressions: {int(total_impr)}')
    if total_impr:
        lines.append(f'  Overall CTR: {total_clicks / total_impr * 100:.2f}%')
    lines.append(f'  Average position: {avg_pos:.2f}')


def add_queries_summary(lines, data_rows):
    if not data_rows:
        return
    total_clicks = sum(float(r[1]) for r in data_rows if r[1])
    total_impr = sum(float(r[2]) for r in data_rows if r[2])
    with_clicks = len([r for r in data_rows if r[1] and float(r[1]) > 0])
    lines.append('')
    lines.append('SUMMARY (Queries):')
    lines.append(f'  Total queries listed: {len(data_rows)}')
    lines.append(f'  Total clicks (all queries): {int(total_clicks)}')
    lines.append(f'  Total impressions (all queries): {int(total_impr)}')
    lines.append(f'  Queries with at least 1 click: {with_clicks}')


def add_pages_summary(lines, data_rows):
    if not data_rows:
        return
    total_clicks = sum(float(r[1]) for r in data_rows if r[1])
    total_impr = sum(float(r[2]) for r in data_rows if r[2])
    lines.append('')
    lines.append('SUMMARY (Pages):')
    lines.append(f'  Total pages listed: {len(data_rows)}')
    lines.append(f'  Total clicks (all pages): {int(total_clicks)}')
    lines.append(f'  Total impressions (all pages): {int(total_impr)}')


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_PATH
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT_PATH

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    lines = [
        '=' * 80,
        'ENDPOINT MEDIA — GOOGLE SEARCH CONSOLE EXPORT REPORT',
        '=' * 80,
        f'Source spreadsheet: {SOURCE_URL}',
        f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        f'Total tabs exported: {len(wb.sheetnames)}',
        f'Tab names: {", ".join(wb.sheetnames)}',
        '',
    ]

    if 'Filters' in wb.sheetnames:
        ws = wb['Filters']
        lines.append('=' * 80)
        lines.append('TAB: FILTERS (Report Parameters)')
        lines.append('=' * 80)
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] is None and (len(row) < 2 or row[1] is None):
                continue
            value = fmt(row[1]) if len(row) > 1 else ''
            lines.append(f'  {fmt(row[0])}: {value}')
        lines.append('')

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Filters':
            continue

        ws = wb[sheet_name]
        lines.append('=' * 80)
        lines.append(f'TAB: {sheet_name.upper()}')
        lines.append('=' * 80)
        lines.append(f'Rows: {ws.max_row} | Columns: {ws.max_column}')
        lines.append('')

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append('(empty sheet)')
            lines.append('')
            continue

        header = [fmt(c) for c in rows[0]]
        data_rows = write_table(lines, rows, header)

        if sheet_name == 'Chart':
            add_chart_summary(lines, data_rows)
        elif sheet_name == 'Queries':
            add_queries_summary(lines, data_rows)
        elif sheet_name == 'Pages':
            add_pages_summary(lines, data_rows)

        lines.append('')

    output_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f'Wrote {output_path} ({output_path.stat().st_size} bytes)')


if __name__ == '__main__':
    main()
